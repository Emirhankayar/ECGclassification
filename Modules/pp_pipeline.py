import csv
import h5py
import shutil
import zipfile
import sklearn
import openpyxl
import constants
import collections
import numpy as np
from tqdm import tqdm
from pathlib import Path


def unzip_file(zip_filepath, output_dir):
    with zipfile.ZipFile(zip_filepath, "r") as zip_ref:
        files = zip_ref.namelist()

        for file in tqdm(files, desc="Unzipping content...", unit="file"):
            zip_ref.extract(file, output_dir)
    print(f"\n Extracted all files to {output_dir}")


def read_xlsx(file_path):
    patient_dict = {}
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    header = {col: idx for idx, col in enumerate(next(ws.iter_rows(values_only=True)))}

    for row in tqdm(
        ws.iter_rows(values_only=True), desc="Reading Excel data...", unit="row"
    ):
        if row[header["FileName"]] and row[header["Rhythm"]]:
            patient_id = row[header["FileName"]]
            rhythm = row[header["Rhythm"]]

            if rhythm in rhythm_mapping:
                mapped_rhythm = rhythm_mapping[rhythm]
                target = label_encoder.transform([mapped_rhythm])[0]
                patient_dict[patient_id] = target

    print(f"\n Total patient entries loaded: {len(patient_dict)}")

    wb.close()
    return patient_dict


def save_patient_hdf5(patient_id, sequences, target_class):
    file_path = constants.DATA_TEMP / f"y_{target_class}" / f"{patient_id}.h5"

    Path(file_path.parent).mkdir(parents=True, exist_ok=True)

    with h5py.File(file_path, "w") as f:
        f.create_dataset("sequences", data=sequences)


def read_csv(directory, patient_dict):
    for file in tqdm(
        Path(directory).glob("*.csv"), desc="Processing CSV files...", unit="file"
    ):
        filename = file.stem

        if filename in patient_dict:
            target_class = patient_dict[filename]
            sequences = []

            with open(file, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)

                for row in reader:
                    try:
                        sequence = list(map(float, row))
                        if len(sequence) == 12:
                            sequences.append(sequence)
                    except ValueError:
                        continue

            if len(sequences) == 5000:
                sequences = np.array(sequences, dtype=np.float32)
                if not np.isnan(sequences).any():
                    save_patient_hdf5(filename, sequences, target_class)


def check_dataset(directory=constants.DATA_TEMP):
    dataset_dir = Path(directory)

    if not dataset_dir.exists():
        print("\n Dataset directory not found!")
        return

    seen_files = collections.defaultdict(list)

    for class_dir in sorted(dataset_dir.iterdir()):
        if class_dir.is_dir():
            class_label = class_dir.stem
            print(f"\n Checking class: {class_label}")

            patient_files = list(class_dir.glob("*.h5"))
            if not patient_files:
                print(f"\n No patient files found in {class_label}")
                continue

            for patient_file in tqdm(
                patient_files, desc=f"Processing {class_label}", unit="file"
            ):
                filename = patient_file.stem
                seen_files[filename].append(class_label)

            first_file = patient_files[0]
            with h5py.File(first_file, "r") as f:
                if "sequences" in f:
                    sequences = f["sequences"][:]

                    if sequences.shape != (5000, 12):
                        print("\n Data shape is incorrect! Expected (5000, 12)")
                else:
                    print(f"\n Missing 'sequences' dataset in {first_file.stem}")

    duplicates = {k: v for k, v in seen_files.items() if len(v) > 1}
    if duplicates:
        print("\n Duplicate Files Found!")
        for filename, classes in duplicates.items():
            print(f" - {filename} appears in: {', '.join(classes)}")
    else:
        print("\n No duplicate filenames found across directories.")


def bin_array(data, window_size=constants.WINDOW_SIZE):
    sequence_length, num_channels = data.shape
    new_sequence_length = sequence_length // window_size
    reshaped = data[: new_sequence_length * window_size].reshape(
        new_sequence_length, window_size, num_channels
    )
    binned_data = np.mean(reshaped, axis=1)
    return binned_data


def pp_dataset(directory=constants.DATA_TEMP, output_directory=constants.DATASET):
    dataset_dir = Path(directory)
    output_dir = Path(output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)

    for split in ["train", "val", "test"]:
        (output_dir / split).mkdir(exist_ok=True)

    file_labels = []

    for class_dir in sorted(dataset_dir.iterdir()):
        if class_dir.is_dir():
            class_label = class_dir.stem
            (output_dir / "train" / class_label).mkdir(parents=True, exist_ok=True)
            (output_dir / "val" / class_label).mkdir(parents=True, exist_ok=True)
            (output_dir / "test" / class_label).mkdir(parents=True, exist_ok=True)

            for patient_file in class_dir.glob("*.h5"):
                file_labels.append((patient_file, class_label))

    files, labels = zip(*file_labels)

    X_train, X_temp, y_train, y_temp = sklearn.model_selection.train_test_split(
        files, labels, test_size=0.2, random_state=42, stratify=labels
    )
    X_val, X_test, y_val, y_test = sklearn.model_selection.train_test_split(
        X_temp, y_temp, test_size=0.25, random_state=42, stratify=y_temp
    )

    def process_and_save(files, labels, split, scaler=None):
        for file, label in tqdm(
            zip(files, labels), desc=f"Processing {split} files", unit="file"
        ):
            with h5py.File(file, "r") as f:
                if "sequences" in f:
                    sequences = f["sequences"][:]
                    sequences = bin_array(sequences)

                    if sequences.shape != (5000 // constants.WINDOW_SIZE, 12):
                        print(
                            f"\n Skipping {file.stem}, incorrect shape: {sequences.shape}"
                        )
                        continue

                    if split == "train":
                        scaler = sklearn.preprocessing.MinMaxScaler()
                        sequences = scaler.fit_transform(sequences)
                    else:
                        sequences = scaler.transform(sequences)

                    save_path = output_dir / split / label / file.name
                    with h5py.File(save_path, "w") as f_out:
                        f_out.create_dataset("sequences", data=sequences)
                else:
                    print(f"\n Missing 'sequences' dataset in {file.stem}")
        return scaler

    scaler = None
    scaler = process_and_save(X_train, y_train, "train", scaler)
    scaler = process_and_save(X_val, y_val, "val", scaler)
    scaler = process_and_save(X_test, y_test, "test", scaler)


def rm_dir(directory: Path):
    if directory.exists() and directory.is_dir():
        print(f"\n Deleting directory: {directory}")
        shutil.rmtree(directory)
    else:
        print(f"\n Directory {directory} not found.")


if __name__ == "__main__":
    print("\n Initializing data preprocessing module...")

    with zipfile.ZipFile(constants.ZIP_PATH, "r") as zip_ref:
        print("\n Extracting...")
        zip_ref.extractall()
        print("\n Extracted all raw data.")

    rhythm_mapping = {
        "AFIB": "AFIB",
        "AF": "AFIB",
        "SVT": "GSVT",
        "AT": "GSVT",
        "SAAWR": "GSVT",
        "ST": "GSVT",
        "AVNRT": "GSVT",
        "AVRT": "GSVT",
        "SB": "SB",
        "SR": "SR",
        "SA": "SR",
    }

    label_encoder = sklearn.preprocessing.LabelEncoder()
    encoded_classes = list(rhythm_mapping.values())
    label_encoder.fit(encoded_classes)

    unzip_file(constants.ZIP_CONTENT, constants.ZIP_CONTENT_OUTPUT)
    patient_dict = read_xlsx(constants.XLSX_PATH)
    read_csv(constants.CSV_PATH, patient_dict)
    check_dataset(constants.DATA_TEMP)
    pp_dataset(constants.DATA_TEMP, constants.DATASET)
    rm_dir(constants.DATA_TEMP)
    print(
        f"\n Preprocessing succesfully finished... \
        \n\n(#_samples,{constants.WINDOW_SIZE},12) is the final data shape.\
        \n\nIf there is a mistake modify the 'WINDOW_SIZE' variable in constants.py."
    )
